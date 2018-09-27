import re, sys
import openpyxl as px

def get_key(r, c, ws):
    key = ws.cell(row = r, column = c).value
    if (key is None) or (key == ""): return ""
    
    # trim \r, \n
    m = re.search(r"^(.*)[\r\n](.*)", key)
    if m: key = m.group(1) + m.group(2)

    # trim '*' from mandatory key like 'User Name *'.
    m = re.match(r"\A(.+) \*\Z", key)
    if m: return m.group(1)
    return key

def get_value(r, c, ws):
    val = ws.cell(row = r, column = c).value
    if (val is None) or (val == ""): return ""
    if val == "-": return ""
    if val == "■": return True
    if val == "□": return False
    if re.match(r"\A\(.+\)\Z", str(val)): return ""
    return val

# scan header line and get the list of target column numbers.
def get_all_target(r, c, ws, sheet_type = "Normal"):
    target_list = []

    max_c = 255  # prevent infinite loop.
    c += 1

    while True:
        key = get_key(r, c, ws)

        if key == "備考" or c >= 255: break

        if sheet_type == "MultiColumn2":
            if target_list:
                target_list[0].append(c)
            elif key:
                # start column of data.
                target_list.append([c])

        if key and (key != "既定値"):
            print("{} at column {} is found".format(key, c), file = sys.stderr)
            if sheet_type == "Normal":
                target_list.append(c)
            elif sheet_type == "MultiColumn":
                target_list.append([])
                target_list[-1].append(c)

        if target_list and (not key) and (sheet_type == "MultiColumn"):
            target_list[-1].append(c)

        c += 1

    return target_list

# get values for specified entry.
def scan_v(r, c, ws, target_column):
    result = {}

    while True:
        key = get_key(r, c, ws)
        if key:
            # trim '# and digit' from repeatable key like 'NSX Edge Appliance #1'.
            m = re.match(r"\A(.+) *#.*\Z", key)
            if m:
                key = m.group(1).rstrip()

                # make list for repeatable keys.
                if key not in result: result[key] = []
            else:
                result[key] = get_value(r, target_column, ws)

            # save current key to store nested keys.
            old_key = key
            r += 1
        else:
            # check nested key exists.
            key = get_key(r, c + 1, ws)
            if not key: break

            r, nested_entries = scan_v(r, c + 1, ws, target_column)
            if isinstance(result[old_key], list):
                result[old_key].append(nested_entries)
            else:
                result[old_key] = nested_entries

    return [r, result]

def extract(ws, start_r = 4, start_c = 2, sheet_type = "Normal"):
    result_list = []

    if sheet_type == "Normal":
        for target_column in get_all_target(start_r, start_c, ws):
            _, result = scan_v(start_r, start_c, ws, target_column)
            result_list.append(result)
    elif (sheet_type == "MultiColumn") or (sheet_type == "MultiColumn2"):
        for target_columns in get_all_target(start_r, start_c, ws, sheet_type):
            result = None
            repeatable_keys = None
            repeatable_items = None
            for target_column in target_columns:
                _, tmp = scan_v(start_r, start_c, ws, target_column)
                if result is None:
                    result = tmp
                    repeatable_keys = get_repeatable_keys(tmp)
                    repeatable_items = tmp
                    for k in repeatable_keys:
                        repeatable_items = repeatable_items[k]
                else:
                    for k in repeatable_keys:
                        tmp = tmp[k]
                    for item in tmp:
                        repeatable_items.append(item)
            result_list.append(result)

    return result_list

def get_repeatable_keys(root):
    if not isinstance(root, dict): return []
    for k in root.keys():
        if isinstance(root[k], list):
            return [k]
        else:
            keys = get_repeatable_keys(root[k])
            if keys:
                result = [k]
                result.extend(keys)
                return result
    return []


if __name__ == '__main__':
    if len(sys.argv) < 2:
         print("usage {} <paramenter_sheet.xlsx>".format(sys.argv[0]))
         exit()

    parameter_sheet = sys.argv[1]
    wb = px.load_workbook(parameter_sheet, data_only = True)
    result_list = extract(wb.get_sheet_by_name('NSX Edge Deploy'))
    wb.close()
    print(result_list)

